from datetime import datetime
import os
import shutil
from qrcode.main import QRCode
from qrcode.constants import ERROR_CORRECT_L
from openpyxl import Workbook
from typing import Optional
import shutil
import zipfile
import os

from redis_om import HashModel, Field
from abc import ABC


def url_from_host_and_tokenid(host, id, protocol="http"):
    return f"{protocol}://{host}/?id={id}"


class Token(HashModel):
    host: str
    description: str
    email: str
    timestamp: Optional[datetime] = Field(index=True)

    def __init__(self, **data):
        if 'timestamp' not in data:
            data['timestamp'] = datetime.now()
        super().__init__(**data)

    def __str__(self):
        return (
            f"Token(host={self.host}, description={
                self.description}, email={self.email}, "
            f"token_type={self.token_type}, timestamp={
                self.timestamp}, id={self.pk})"
        )

    def url(self):
        return url_from_host_and_tokenid(self.host, self.pk)

    def write_out(self) -> None:
        pass


class URLToken(Token):
    token_type: str = Field(default="URLToken", index=True)

    def write_out(self):
        with open(
            f"honeytokens/URL_{self.description}", "w"
        ) as output_file:
            output_file.write(self.url())

    def filename(self):
        return f"URL_{self.description}.txt"

    def __str__(self):
        return (
            f"URLToken(host={self.host}, description={
                self.description}, email={self.email}, "
            f"token_type={self.token_type}, timestamp={self.timestamp}, id={
                self.pk}, url={url_from_host_and_tokenid(self.host, self.pk)}"
        )


class QRToken(Token):
    token_type: str=Field(default="QRToken", index=True)

    def write_out(self):
        self.create_qr_code(f"honeytokens/{self.filename()}")

    def create_qr_code(self, file_name: str):
        # Create a QR code object
        qr=QRCode(
            version=1,
            error_correction=ERROR_CORRECT_L,
            box_size=10,
            border=4,
        )
        # Add data to the QR code
        qr.add_data(self.url())
        qr.make(fit=True)
        # Create an image from the QR code object
        img=qr.make_image(fill_color="black", back_color="white")
        # Save the image to a file
        img.save(file_name)

    def filename(self):
        return f"QR_{self.description}.jpg"


class ExcelToken(Token):
    token_type: str=Field(default="ExcelToken", index=True)



    def write_out(self):
        filepath=f"honeytokens/{self.filename()}"
        template=f"templates/template.xlsx"
        shutil.copy(template, filepath)
        url=url_from_host_and_tokenid(self.host, self.pk)

        try:
            extracted_dir=f"honeytokens/extracted"
            with zipfile.ZipFile(filepath, 'r') as zip_ref:
                zip_ref.extractall(extracted_dir)
            rels_dir=os.path.join(extracted_dir, 'xl',
                                  'drawings', '_rels', "drawing1.xml.rels")
            with open(rels_dir, "r") as fd:
                contents=fd.read()
            with open(rels_dir, "w") as fd:
                contents=contents.replace(
                    "HONEYDROP_TOKEN_URL", url)
                fd.write(contents)
            new_zip_file=f"honeytokens/modified_{self.filename()}"
            with zipfile.ZipFile(new_zip_file, 'w') as zipf:
                for root, _, files in os.walk(extracted_dir):
                    for file in files:
                        file_path=os.path.join(root, file)
                        zipf.write(file_path, os.path.relpath(
                            file_path, extracted_dir))
            shutil.rmtree(extracted_dir)
            # shutil.rmtree(filepath)


            # shutil.rmtree(extracted_dir)
        except Exception as e:
            print(f"Error modifying zip file: {e}")

        # filepath=f"honeytokens/{self.filename()}"
        # wb=Workbook()
        # ws=wb.active
        # ws['A1']="This is a test for Excel Token."
        # ws['A1'].style='Title'
        # wb.save(filepath)

        # url=url_from_host_and_tokenid(self.host, self.pk)
        # modified_file=f"honeytokens/modified_{self.filename()}"

        # try:
        #     shutil.copy(filepath, modified_file)
        #     extracted_dir=f"honeytokens/extracted_{self.filename()}"
        #     with zipfile.ZipFile(filepath, 'r') as zip_ref:
        #         zip_ref.extractall(extracted_dir)

        #     # Create the drawings folder and the drawing XML
        #     drawings_dir=os.path.join(extracted_dir, 'xl', 'drawings')
        #     os.makedirs(drawings_dir, exist_ok=True)
        #     drawing_xml_path=os.path.join(drawings_dir, 'drawing1.xml')
        #     drawing_xml_content=f"""
        #     <xdr:wsDr xmlns:xdr="http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing" xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main">
        #         <xdr:twoCellAnchor editAs="oneCell">
        #             <xdr:from>
        #                 <xdr:col>61</xdr:col>
        #                 <xdr:colOff>0</xdr:colOff>
        #                 <xdr:row>376</xdr:row>
        #                 <xdr:rowOff>0</xdr:rowOff>
        #             </xdr:from>
        #             <xdr:to>
        #                 <xdr:col>70</xdr:col>
        #                 <xdr:colOff>190500</xdr:colOff>
        #                 <xdr:row>397</xdr:row>
        #                 <xdr:rowOff>177800</xdr:rowOff>
        #             </xdr:to>
        #             <xdr:pic>
        #                 <xdr:nvPicPr>
        #                     <xdr:cNvPr id="3" name="Picture 2">
        #                         <a:extLst>
        #                             <a:ext uri="{{FF2B5EF4-FFF2-40B4-BE49-F238E27FC236}}">
        #                                 <a16:creationId xmlns:a16="http://schemas.microsoft.com/office/drawing/2014/main" id="{{E90E5A81-5E9B-744C-9F18-59568A01D25B}}"/>
        #                             </a:ext>
        #                         </a:extLst>
        #                     </xdr:cNvPr>
        #                     <xdr:cNvPicPr>
        #                         <a:picLocks noChangeAspect="1"/>
        #                     </xdr:cNvPicPr>
        #                 </xdr:nvPicPr>
        #                 <xdr:blipFill>
        #                     <a:blip xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships" r:link="rId1"/>
        #                     <a:stretch>
        #                         <a:fillRect/>
        #                     </a:stretch>
        #                 </xdr:blipFill>
        #                 <xdr:spPr>
        #                     <a:xfrm>
        #                         <a:off x="50355500" y="76403200"/>
        #                         <a:ext cx="7620000" cy="4445000"/>
        #                     </a:xfrm>
        #                     <a:prstGeom prst="rect">
        #                         <a:avLst/>
        #                     </a:prstGeom>
        #                 </xdr:spPr>
        #             </xdr:pic>
        #             <xdr:clientData/>
        #         </xdr:twoCellAnchor>
        #     </xdr:wsDr>
        #     """
        #     with open(drawing_xml_path, 'w', encoding='utf-8') as drawing_file:
        #         drawing_file.write(drawing_xml_content)

        #     # Create the .rels file to link the drawing
        #     rels_dir=os.path.join(extracted_dir, 'xl', 'drawings', '_rels')
        #     os.makedirs(rels_dir, exist_ok=True)
        #     rels_xml_path=os.path.join(rels_dir, 'drawing1.xml.rels')
        #     rels_xml_content=f"""
        #     <Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
        #         <Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/image" Target="{url}" TargetMode="External"/>
        #     </Relationships>
        #     """
        #     with open(rels_xml_path, 'w', encoding='utf-8') as rels_file:
        #         rels_file.write(rels_xml_content)

        #     # Create the sheet1.xml.rels file to link the drawing to the worksheet
        #     sheet_rels_dir=os.path.join(
        #         extracted_dir, 'xl', 'worksheets', '_rels')
        #     os.makedirs(sheet_rels_dir, exist_ok=True)
        #     sheet_rels_xml_path=os.path.join(sheet_rels_dir, 'sheet1.xml.rels')
        #     sheet_rels_xml_content="""
        #     <Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
        #         <Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/drawing" Target="../drawings/drawing1.xml"/>
        #     </Relationships>
        #     """
        #     with open(sheet_rels_xml_path, 'w', encoding='utf-8') as sheet_rels_file:
        #         sheet_rels_file.write(sheet_rels_xml_content)

        #     new_zip_file=f"honeytokens/modified_{self.filename()}"
        #     with zipfile.ZipFile(new_zip_file, 'w') as zipf:
        #         for root, _, files in os.walk(extracted_dir):
        #             for file in files:
        #                 file_path=os.path.join(root, file)
        #                 zipf.write(file_path, os.path.relpath(
        #                     file_path, extracted_dir))

        #     # Optionally, delete the extracted directory
        #     shutil.rmtree(extracted_dir)

        # except Exception as e:
        #     print(f"Error modifying zip file: {e}")
    # def write_out(self):
    #     filepath=f"honeytokens/{self.filename()}"
    #     wb=Workbook()
    #     ws=wb.active
    #     ws['A1']="This is a test for Excel Token."
    #     ws['A1'].style='Title'
    #     wb.save(filepath)
    #     old_string='"http://schemas.openxmlformats.org/spreadsheetml/2006/main"'
    #     url=f'"{url_from_host_and_tokenid(self.host, self.pk)}"'
    #     modified_file=f"honeytokens/modified_{self.filename()}"
    #     try:
    #         shutil.copy(filepath, modified_file)
    #         extracted_dir=f"honeytokens/extracted_{self.filename()}"
    #         with zipfile.ZipFile(filepath, 'r') as zip_ref:

    #             zip_ref.extractall(extracted_dir)

    #         styles_path=os.path.join(
    #             extracted_dir, 'xl', 'styles.xml')
    #         print("the path of the file is: ", styles_path)
    #         with open(styles_path, 'r', encoding='utf-8') as file:
    #             styles_content=file.read()

    #         new_styles_content=styles_content.replace(old_string, url)
    #         with open(styles_path, 'w', encoding='utf-8') as file:
    #             file.write(new_styles_content)
    #         new_zip_file=f"honeytokens/modified_{self.filename()}"

    #         with zipfile.ZipFile(new_zip_file, 'w') as zipf:
    #             for root, _, files in os.walk(extracted_dir):
    #                 for file in files:
    #                     file_path=os.path.join(root, file)
    #                     zipf.write(file_path, os.path.relpath(
    #                         file_path, extracted_dir))

    #     # Toggle comment to erase the extracted dir
    #         # shutil.rmtree(extracted_dir)

    #     except Exception as e:
    #         print(f"Error modifying zip file: {e}")

    def filename(self):
        return f"Excel_{self.description}.xlsx"

    def check_modified_xml(self, file_path):
        print('Checking modified XML content...')
        old_string='xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"'
        url=url_from_host_and_tokenid(self.host, self.pk)
        extracted_dir=f"honeytokens/extracted_{self.filename}"
        with zipfile.ZipFile(file_path, 'r') as zip_ref:
            zip_ref.extractall(extracted_dir)

        styles_path=os.path.join(extracted_dir, 'xl', 'styles.xml')
        with open(styles_path, 'r', encoding='utf-8') as file:
            styles_content=file.read()
            print("the content of the file is: ", styles_content)
            if old_string in styles_content:
                print('URL not inserted in the XML content.')
            elif url in styles_content:
                print('URL inserted in the XML content.')
            else:
                print('URL not found in the XML content.')
            print('Modified XML content checked successfully.')


class DockerfileToken(Token):
    token_type: str=Field(default="DockerfileToken", index=True)

    def write_out(self):
        dockerfile_payload=self.get_dockerfile_payload()
        with open(
            f"honeytokens/{self.filename()}", "w"
        ) as output_file:

            output_file.write(dockerfile_payload)

        print(f"Your Dockerfile Token payload: {dockerfile_payload}")

    def get_dockerfile_payload(self):
        payload=f"""
CMD ["bash", "-c",
    "echo -e 'GET /?id={self.pk} HTTP/1.1\\r\\nHost: {self.host}\\r\\nConnection: close\\r\\n\\r\\n' >/dev/tcp/{self.host}/5000"]
"""
        return payload

    def filename(self):
        return f"Dockerfile_{self.description}.txt"


class WindowsDirectoryToken(Token):
    token_type: str=Field(default="WindowsDirectoryToken", index=True)

    def write_out(self):
        """
        Creamos una carpeta con un archivo desktop.ini que tiene un un icono con la url del token.
        Cuando el directorio se abre en el Explorador de Windows, el sistema intenta acceder al icono,
        y se hace el request al servidor.
        """

        icon_url=url_from_host_and_tokenid(self.host, self.pk)
        desktop_ini_content=f"""
        [.ShellClassInfo]
        IconResource={icon_url},0
        """
        with ZipFile(self.filename(), "w") as zip_file:
            zip_file.writestr("desktop.ini", desktop_ini_content.strip())

    def filename(self):
        return f"honeytokens/WindowsDirectory_{self.description}.zip"


class HTMLToken(Token):
    token_type: str=Field(default="HTMLToken", index=True)
    allowed_url: str
    input_html_path: str

    def write_out(self):
        tokenized_html=self.tokenize_html(self.input_html_path)
        with open(f"honeytokens/{self.filename()}", "w") as html_file:
            html_file.write(tokenized_html)

    def tokenize_html(self, html_file_path: str) -> str:
        with open(html_file_path, "r") as file:
            html_content=file.read()

        alert_script=f"""
<script>
if (window.location.hostname !== "{self.allowed_url}") {{
    fetch("{url_from_host_and_tokenid(self.host, self.pk)}");
}}
</script>
"""
        return html_content + alert_script

    def filename(self):
        return f"HTML_{self.description}.html"
